# journal_ledger.py

from journal_file import Journal
from openpyxl import Workbook

def ledger_category(journal_name="분개장_샘플.xlsx", file_name="계정별원장.xlsx"):
    # 분개장 데이터 로딩
    jnl = Journal(journal_name)

    # 워크북 생성
    wb = Workbook()

    # 기본적으로 생성되는 활성 시트 제거
    default_sheet = wb.active
    wb.remove(default_sheet)


    # 각 카테고리에 대해 시트 생성 및 데이터 쓰기
    for category in jnl.df["계정과목"].unique():
        # category 이름으로 신규 시트 생성
        ws = wb.create_sheet(title=category)

        # 시트에 첫번째 셀에 category 이름 출력
        ws.append([category])
        #ws.cell(row=1, column=1, value=category)

        # 한칸 띄우기
        ws.append([])

        # 컬럼명 출력
        ws.append(tuple(jnl.df.columns))

        # category에 해당하는 계정과목 DataFrame에 대하여 한행씩 row 변수에 대입
        for r_idx, row in enumerate(jnl.계정과목(category).df.itertuples(index=False), start=4):
            # itertuples 메서드 대신 iterrows 메서드를 사용해도 되며 차이점은 반복 속도임
            # (itertuples 메서드가 iterrows 메서드 보다 빠르게 작동)

            # 한 행(row)에 대하여 컬럼번호(c_idx)를 증가시키면서 한 셀씩 출력
            for c_idx, value in enumerate(row, start=1):
                ws.cell(row=r_idx, column=c_idx, value=value)

    # 최종 완성된 파일 엑셀 저장
    wb.save(file_name)


def ledger_client(journal_name="분개장_샘플.xlsx", file_name="거래처원장.xlsx"):
    # 분개장 데이터 로딩
    jnl = Journal(journal_name)
    
    # 워크북 생성
    wb = Workbook()

    # 기본적으로 생성되는 활성 시트 제거
    default_sheet = wb.active
    wb.remove(default_sheet)


    # 각 카테고리에 대해 시트 생성 및 데이터 쓰기
    for client in jnl.df["거래처명"].unique():
        # client 이름으로 신규 시트 생성
        ws = wb.create_sheet(title=client)

        # 시트에 첫번째 셀에 client 이름 출력
        ws.append([client])
        #ws.cell(row=1, column=1, value=client)

        # 한칸 띄우기
        ws.append([])

        # 컬럼명 출력
        ws.append(tuple(jnl.df.columns))

        # client에 해당하는 거래처명 DataFrame에 대하여 한행씩 row 변수에 대입
        for r_idx, row in enumerate(jnl.거래처명(client).df.itertuples(index=False), start=4):
            # itertuples 메서드 대신 iterrows 메서드를 사용해도 되며 차이점은 반복 속도임
            # (itertuples 메서드가 iterrows 메서드 보다 빠르게 작동)

            # 한 행(row)에 대하여 컬럼번호(c_idx)를 증가시키면서 한 셀씩 출력
            for c_idx, value in enumerate(row, start=1):
                ws.cell(row=r_idx, column=c_idx, value=value)

    # 최종 완성된 파일 엑셀 저장
    wb.save(file_name)