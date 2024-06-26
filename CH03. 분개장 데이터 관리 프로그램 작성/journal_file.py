import pandas as pd

pd.set_option('display.max_rows', 20)
pd.set_option('display.max_columns', 100)
pd.set_option('display.max_colwidth', 50)
pd.set_option('display.width', 300)
pd.set_option('display.expand_frame_repr', True)

from pandas import DataFrame, Series
from datetime import datetime, date
from openpyxl import load_workbook


def load_xlfile(filename):
    colname_row_number = 4
    start_row_number = 5
    last_row_value = "합 계"


    wb = load_workbook(filename)
    ws = wb.active

    colnames = [x.value for x in ws[colname_row_number]]

    datalist = []
    for row in ws.iter_rows(min_row=start_row_number, max_col=len(colnames), values_only=True):
        if row[0] == last_row_value:
            break
        row = list(row)
        
        if row[0] is not None:
            일자 = row[0].split('/')
            일자 = date(2024, int(일자[0]), int(일자[1]))
            전표번호 = row[1]    
        
        row[0] = 일자
        row[1] = 전표번호
        if pd.isna(row[5]):
            row[5] = 0
        if pd.isna(row[6]):
            row[6] = 0

        datalist.append(row)
    journal = DataFrame(datalist, columns=colnames)

    category_table = pd.DataFrame([
            ['자본금', '자본', '자본금', '자본금'],
            ['보통예금', '유동자산', '현금및현금성자산', '보통예금'],
            ['객실수입', '매출', '매출', '객실수입'],
            ['고객용품비', '판매비와관리비', '고객용품비', '고객용품비'],
            ['기타영업비용', '판매비와관리비', '판매비용', '기타영업비용'],
            ['판매수수료', '판매비와관리비', '판매비용', '판매수수료'],
            ['세금과공과', '판매비와관리비', '기타판관비', '세금과공과'],
            ['수도광열비', '판매비와관리비', '기타판관비', '수도광열비'],
            ['급여', '판매비와관리비', '인건비', '급여'],
            ['노무용역비', '판매비와관리비', '인건비', '노무용역비'],
            ['이자수입', '영업외수익', '영업외수익', '이자수입']
        ],
        columns=['구분', '대분류', '중분류', '소분류']
    )
    category_table.set_index('구분', inplace=True)

    journal['대분류'] = None
    journal['중분류'] = None
    journal['소분류'] = None

    for idx in journal.index:
        # journal 데이터프레임의 "계정과목" 컬럼 값 추출
        key = journal.loc[idx, '계정과목']

        # category_table에서 추출된 키값에 해당하는 "대분류", "중분류", "소분류" 값을 journal에 대체
        journal.loc[idx, '대분류'] = category_table.loc[key, '대분류']
        journal.loc[idx, '중분류'] = category_table.loc[key, '중분류']
        journal.loc[idx, '소분류'] = category_table.loc[key, '소분류']

    journal = journal[
        [
            '일자', '전표번호', '대분류', '중분류', '소분류', '계정코드', 
            '계정과목', '적요', '차변', '대변', '구분', '거래처명'
            ]
        ]
    
    return journal


class Journal:
    def __init__(self, input_value):
        if type(input_value) is DataFrame:
            # input_value의 형식이 DataFrame인 경우 바로 df 변수에 저장
            self.df = input_value
        elif type(input_value) is str and input_value[-4:] == "xlsx":
            # input_value의 형식이 string이고, 마지막 4개의 값이 "xlsx"인 경우
            # load_xlfile 함수를 이용하여 엑셀 파일의 데이터를 가져와 df 변수에 저장
            self.df = load_xlfile(input_value)
        else:
            raise ValueError("Error: 데이터 프레임 또는 엑셀파일명을 입력하시요.")
        
        self.전표번호 = self.Filter(self, "전표번호")
        self.대분류 = self.Filter(self, "대분류")
        self.중분류 = self.Filter(self, "중분류")
        self.소분류 = self.Filter(self, "소분류")
        self.계정코드 = self.Filter(self, "계정코드")
        self.계정과목 = self.Filter(self, "계정과목")
        self.거래처명 = self.Filter(self, "거래처명")

    class Filter:
        def __init__(self, parent_instance, attr_name):
            self.parent_instance = parent_instance
            self.attr_name = attr_name

        def __call__(self, lookup_value):
            df = self.parent_instance.df
            filtered_df = df[df[self.attr_name] == lookup_value]
            return Journal(filtered_df)