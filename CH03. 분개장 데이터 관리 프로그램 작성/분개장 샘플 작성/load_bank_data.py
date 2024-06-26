import pandas as pd
from pandas import DataFrame, Series
import numpy as np
import openpyxl as xl
from cafle.genfunc import print_rounding

#Functions
def isintuple(lst, tpl):
    for val in lst:
        if val in tpl:
            return True
    return False

def change_colname(df, name_from, name_to):
    if name_from in df.columns:
        df.rename(columns={name_from: name_to}, inplace=True)
    return None

def del_bracket(val_str):
    val_str = val_str.replace(")", "")
    val_str = val_str.replace("(", "")
    return val_str

def sort_columns(df):
    cols_name = Series(["거래일", "거래일시", "입금", "출금", "잔액", "적요", "내용", "상대계좌예금주명", "상대은행", "상대계좌번호"])
    col1 = cols_name[cols_name.isin(df.columns)]
    col2 = df.columns[Series(df.columns.isin(cols_name)).map(lambda x: not x)]
    cols_name_final = pd.concat([col1, Series(col2)])
    return df[cols_name_final]

def read_bnkdata(ws):
    bnk = {}

    #Read bank overview data
    ov = {}
    for row in ws.iter_rows(min_row=1, max_col=2, max_row=9, values_only=True):
        if pd.isna(row[0]):
            continue
        ov[row[0]] = row[1]
    bnk['ov'] = ov

    #Read values to DataFrame
    tmpidx = [x.value for x in ws[10]]
    colidx = []
    for val in tmpidx:
        if pd.isna(val):
            continue
        colidx.append(val)

    arrtmp = []
    noneNo = 0
    for row in ws.iter_rows(min_row=11, max_col=len(colidx), max_row=50000, values_only=True):
        if isintuple(["거래일시", "합계", "합   계"], row):
            pass
        elif all(pd.isna(list(row))):
            noneNo += 1
            if noneNo == 10:
                break
            pass
        else:
            arrtmp.append(row)
    df = DataFrame(arrtmp, columns=colidx)

    #Edit dataframe column names
    change_colname(df, '입금액', '입금')
    change_colname(df, '출금액', '출금')
    change_colname(df, '입금(원)', '입금')
    change_colname(df, '지급(원)', '출금')
    change_colname(df, '거래후 잔액', '잔액')
    change_colname(df, '거래후 잔액(원)', '잔액')
    change_colname(df, '거래후잔액', '잔액')
    change_colname(df, '출금통장표시내용', '내용')
    change_colname(df, '거래내용', '내용')

    #Edit dataframe columns
    if '거래일시' in df.columns:
        df['거래일시'] = df['거래일시'].map(del_bracket)
        df['거래일시'] = pd.to_datetime(df['거래일시'])
        df['거래일'] = df['거래일시'].map(lambda x: x.date())

    bnk['df'] = sort_columns(df)

    return bnk

def load_xlfile(filename):
    #Load excel file
    wb = xl.load_workbook(filename)
    bnks = {}
    for ws in wb.worksheets:
        bnks[ws.title] = read_bnkdata(ws)
    return bnks

class Bnks:
    def __init__(self):
        self.filename = "계좌_거래내역/은행계좌내역_통합.xlsx"
        self.bnks = load_xlfile(self.filename)

    @property
    def ov(self):
        rslt = {key: item['ov'] for key, item in self.bnks.items()}
        return rslt

    @property
    def df(self):
        rslt = {key: item['df'] for key, item in self.bnks.items()}
        return rslt

    def valondate(self, valdate):
        lstdf = []
        for key, item in self.bnks.items():
            df = item['df']
            dfdt = df[df['거래일'] == valdate].copy()
            dfcolname = ['계좌명']
            dfcolname.extend(dfdt.columns)

            dfdt['계좌명'] = [key] * len(dfdt)
            dfdt = dfdt[dfcolname]

            lstdf.append(dfdt)
        return pd.concat(lstdf)